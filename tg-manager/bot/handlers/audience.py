"""Audience collection, stats, comparison, CSV export/XLSX export, and user management."""

from __future__ import annotations
import csv
import io
import asyncpg
import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from bot.callbacks import AudCb, BotCb
from bot.keyboards import (
    audience_menu,
    bots_pick,
    back_to_bot,
    user_profile_menu,
    subscription_locked_markup,
)
from bot.states import SendToUser
from bot.utils.subscription import require_plan, locked_text
from database import db
from services import bot_api

router = Router()


@router.callback_query(AudCb.filter(F.action == "menu"))
async def cb_aud_menu(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer()
    count = await db.get_audience_count(pool, row["bot_id"])
    label = f"@{row['username']}" if row["username"] else row["first_name"]
    await callback.message.edit_text(
        f"👥 <b>Аудитория — {label}</b>\n\n"
        "📌 <b>Что это?</b>\n"
        "Здесь хранится список всех людей, которые когда-либо писали вашему боту. Вы можете посмотреть кто они, сравнить с другими ботами, скачать список или написать конкретному пользователю.\n\n"
        "💡 <b>Что можно делать:</b>\n"
        "• <b>Обновить</b> — загрузить новых пользователей из Telegram\n"
        "• <b>Сравнить</b> — посмотреть, сколько пользователей есть у двух ботов одновременно\n"
        "• <b>Экспорт</b> — скачать список пользователей в файл\n"
        "• <b>Профиль</b> — найти и написать конкретному человеку\n\n"
        f"Активных пользователей: <b>{count}</b>",
        parse_mode="HTML",
        reply_markup=audience_menu(row["bot_id"]),
    )


@router.callback_query(AudCb.filter(F.action == "refresh"))
async def cb_refresh(
    callback: CallbackQuery,
    callback_data: AudCb,
    pool: asyncpg.Pool,
    http: aiohttp.ClientSession,
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer()

    await callback.message.edit_text("⏳ Собираю обновления…")

    updates = await bot_api.fetch_updates(http, row["token"])
    users = bot_api.extract_users_from_updates(updates)
    new_count = await db.upsert_users(pool, row["bot_id"], users)
    total = await db.get_audience_count(pool, row["bot_id"])
    label = f"@{row['username']}" if row["username"] else row["first_name"]

    await callback.message.edit_text(
        f"👥 <b>Аудитория {label}</b>\n\n"
        f"Получено апдейтов: {len(updates)}\n"
        f"Новых пользователей: <b>+{new_count}</b>\n"
        f"Всего активных: <b>{total}</b>",
        parse_mode="HTML",
        reply_markup=audience_menu(row["bot_id"]),
    )


@router.callback_query(AudCb.filter(F.action == "stats"))
async def cb_stats(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer()

    stats = await db.get_audience_stats(pool, row["bot_id"])
    label = f"@{row['username']}" if row["username"] else row["first_name"]

    lang_lines = (
        "\n".join(
            f"  <code>{l['lang']}</code>: {l['count']}" for l in stats["languages"]
        )
        or "  нет данных"
    )

    total_all = stats["total"] + stats["inactive"]
    block_pct = round(stats["inactive"] / total_all * 100, 1) if total_all else 0

    daily = await db.get_audience_daily_growth(pool, row["bot_id"], days=7)
    max_day = max((d["count"] for d in daily), default=1)
    graph_lines = []
    for d in daily:
        bar_len = max(1, round(d["count"] * 10 / max_day)) if max_day else 0
        bar = "█" * bar_len
        date_str = d["date"].strftime("%d.%m")
        graph_lines.append(f"  {date_str}: {bar} +{d['count']}")
    graph = "\n".join(graph_lines) if graph_lines else "  нет данных"

    text = (
        f"📊 <b>Статистика аудитории {label}</b>\n\n"
        f"👤 Активных: <b>{stats['total']}</b>\n"
        f"🚫 Заблокировали бота: <b>{stats['inactive']}</b> ({block_pct}%)\n"
        f"📌 Всего за всё время: <b>{total_all}</b>\n\n"
        f"📈 <b>Прирост:</b>\n"
        f"  За сутки: <b>+{stats['joined_today']}</b>\n"
        f"  За 7 дней: <b>+{stats['joined_week']}</b>\n"
        f"  За 30 дней: <b>+{stats['joined_month']}</b>\n\n"
        f"📅 <b>График (7 дней):</b>\n<code>{graph}</code>\n\n"
        f"🌍 <b>Языки (топ-10):</b>\n{lang_lines}"
    )
    await callback.message.edit_text(
        text, parse_mode="HTML", reply_markup=audience_menu(row["bot_id"])
    )


@router.callback_query(AudCb.filter(F.action == "export"))
async def cb_export(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer("⏳ Генерирую CSV…")

    rows = await db.get_audience_full(pool, row["bot_id"])
    if not rows:
        await callback.message.answer("📤 Аудитория пуста — нечего экспортировать.")
        return

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "user_id",
            "username",
            "first_name",
            "last_name",
            "language_code",
            "first_seen",
            "last_seen",
            "is_active",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r["user_id"],
                r["username"] or "",
                r["first_name"] or "",
                r["last_name"] or "",
                r["language_code"] or "",
                r["first_seen"].strftime("%Y-%m-%d %H:%M:%S"),
                r["last_seen"].strftime("%Y-%m-%d %H:%M:%S"),
                r["is_active"],
            ]
        )

    label = f"@{row['username']}" if row["username"] else row["first_name"]
    safe_label = row["username"] or str(row["bot_id"])
    filename = f"audience_{safe_label}.csv"
    content = buf.getvalue().encode("utf-8-sig")

    await callback.message.answer_document(
        BufferedInputFile(content, filename=filename),
        caption=f"📤 Аудитория <b>{label}</b> — {len(rows)} записей",
        parse_mode="HTML",
    )


@router.callback_query(AudCb.filter(F.action == "export_xlsx"))
async def cb_export_xlsx(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer("⏳ Генерирую Excel…")

    rows = await db.get_audience_full(pool, row["bot_id"])
    stats = await db.get_audience_stats(pool, row["bot_id"])

    label = f"@{row['username']}" if row["username"] else row["first_name"]
    safe_label = row["username"] or str(row["bot_id"])

    wb = Workbook()
    # ── Sheet 1: Аудитория ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Аудитория"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "user_id",
        "username",
        "first_name",
        "last_name",
        "language_code",
        "first_seen",
        "last_seen",
        "is_active",
    ]
    col_widths = [14, 22, 20, 20, 14, 22, 22, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    for r_idx, r in enumerate(rows, start=2):
        data = [
            r["user_id"],
            r["username"] or "",
            r["first_name"] or "",
            r["last_name"] or "",
            r["language_code"] or "",
            r["first_seen"].strftime("%Y-%m-%d %H:%M") if r["first_seen"] else "",
            r["last_seen"].strftime("%Y-%m-%d %H:%M") if r["last_seen"] else "",
            "Да" if r["is_active"] else "Нет",
        ]
        row_fill = alt_fill if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    # ── Sheet 2: Статистика ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Статистика")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    stat_header_fill = PatternFill("solid", fgColor="059669")
    ws2.cell(row=1, column=1, value="Показатель").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=1, column=1).fill = stat_header_fill
    ws2.cell(row=1, column=2, value="Значение").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=1, column=2).fill = stat_header_fill

    stat_rows = [
        ("Бот", label),
        ("Всего активных", stats.get("total", 0)),
        ("Неактивных (отписались)", stats.get("inactive", 0)),
        ("Новых за 24ч", stats.get("joined_today", 0)),
        ("Новых за 7 дней", stats.get("joined_week", 0)),
        ("Новых за 30 дней", stats.get("joined_month", 0)),
    ]
    for s_idx, (k, v) in enumerate(stat_rows, start=2):
        ws2.cell(row=s_idx, column=1, value=k).border = border
        ws2.cell(row=s_idx, column=2, value=v).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"audience_{safe_label}.xlsx"

    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"📊 <b>Аудитория {label}</b>\n{len(rows)} пользователей · Excel-файл с форматированием",
        parse_mode="HTML",
    )


@router.callback_query(AudCb.filter(F.action == "scan"))
async def cb_scan(
    callback: CallbackQuery,
    callback_data: AudCb,
    pool: asyncpg.Pool,
    http: aiohttp.ClientSession,
) -> None:

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer()

    await callback.message.edit_text("⚡ Сканирую все доступные апдейты…")

    from database import db as _db
    from services import bot_api as _api

    start_offset = await _db.get_update_offset(pool, callback_data.bot_id)
    users, last_id = await _api.scan_all_users(
        http, row["token"], start_offset=start_offset
    )

    new_count = 0
    if users:
        new_count = await db.upsert_users(pool, row["bot_id"], users)
    if last_id > start_offset:
        await db.set_update_offset(pool, callback_data.bot_id, last_id)

    total = await db.get_audience_count(pool, row["bot_id"])
    label = f"@{row['username']}" if row["username"] else row["first_name"]
    await callback.message.edit_text(
        f"👥 <b>Аудитория {label}</b>\n\n"
        f"⚡ Просканировано апдейтов до ID #{last_id}\n"
        f"Найдено уникальных пользователей: <b>{len(users)}</b>\n"
        f"Новых добавлено: <b>+{new_count}</b>\n"
        f"Всего активных: <b>{total}</b>",
        parse_mode="HTML",
        reply_markup=audience_menu(row["bot_id"]),
    )


@router.callback_query(AudCb.filter(F.action == "compare"))
async def cb_compare_pick(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    bots = await db.get_bots(pool, callback.from_user.id)
    others = [b for b in bots if b["bot_id"] != callback_data.bot_id]
    if not others:
        await callback.answer(
            "Нужен хотя бы ещё один бот для сравнения.", show_alert=True
        )
        return
    await callback.answer()
    await callback.message.edit_text(
        "⚖️ Выберите второй бот для сравнения аудиторий:",
        reply_markup=bots_pick(bots, exclude_bot_id=callback_data.bot_id),
    )


@router.callback_query(AudCb.filter(F.action == "pick_b"))
async def cb_compare_result(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    row_a = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    row_b = await db.get_bot(pool, callback_data.target_id, callback.from_user.id)
    if not row_a or not row_b:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await callback.answer()

    stats = await db.compare_audiences(pool, row_a["bot_id"], row_b["bot_id"])
    label_a = f"@{row_a['username']}" if row_a["username"] else row_a["first_name"]
    label_b = f"@{row_b['username']}" if row_b["username"] else row_b["first_name"]

    await callback.message.edit_text(
        f"⚖️ <b>Сравнение аудиторий</b>\n\n"
        f"<b>{label_a}</b>: {stats['count_a']} чел.\n"
        f"<b>{label_b}</b>: {stats['count_b']} чел.\n\n"
        f"🔁 Пересечение: <b>{stats['overlap']}</b> чел.\n"
        f"   {stats['overlap_pct_a']}% от {label_a}\n"
        f"   {stats['overlap_pct_b']}% от {label_b}",
        parse_mode="HTML",
        reply_markup=back_to_bot(row_a["bot_id"]),
    )


# ── Send to specific user ──────────────────────────────────────────────────


@router.callback_query(AudCb.filter(F.action == "send_user"))
async def cb_send_user(
    callback: CallbackQuery, callback_data: AudCb, state: FSMContext
) -> None:
    from aiogram.utils.keyboard import InlineKeyboardBuilder as _Kb

    await callback.answer()
    await state.set_state(SendToUser.waiting_user_id)
    await state.update_data(bot_id=callback_data.bot_id)
    kb = _Kb()
    kb.button(
        text="❌ Отмена",
        callback_data=AudCb(action="menu", bot_id=callback_data.bot_id),
    )
    await callback.message.edit_text(
        "📤 <b>Написать пользователю</b>\n\n"
        "Введите Telegram User ID пользователя\n"
        "(число, например <code>123456789</code>):",
        parse_mode="HTML",
        reply_markup=kb.as_markup(),
    )


@router.message(SendToUser.waiting_user_id, F.text)
async def msg_send_user_id(message: Message, state: FSMContext) -> None:
    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Неверный формат. Введите числовой User ID:")
        return
    await state.update_data(target_user_id=user_id)
    await state.set_state(SendToUser.waiting_message)
    await message.answer(
        f"✅ User ID: <code>{user_id}</code>\n\nТеперь введите текст сообщения:",
        parse_mode="HTML",
    )


@router.message(SendToUser.waiting_message, F.text)
async def msg_send_user_text(
    message: Message, state: FSMContext, pool: asyncpg.Pool, http: aiohttp.ClientSession
) -> None:
    data = await state.get_data()
    await state.clear()
    bot_id = data["bot_id"]
    target_user_id = data["target_user_id"]

    row = await db.get_bot(pool, bot_id, message.from_user.id)
    if not row:
        await message.answer("Бот не найден.")
        return

    ok, retry = await bot_api.send_message(
        http, row["token"], target_user_id, message.text
    )
    if ok:
        await message.answer(
            f"✅ Сообщение доставлено пользователю <code>{target_user_id}</code>.",
            parse_mode="HTML",
            reply_markup=audience_menu(bot_id),
        )
    else:
        await message.answer(
            f"❌ Не удалось отправить. Пользователь <code>{target_user_id}</code> "
            "мог заблокировать бота или не начинал с ним диалог.",
            parse_mode="HTML",
            reply_markup=audience_menu(bot_id),
        )


# ── Export audience from bot menu ─────────────────────────────────────────


@router.callback_query(BotCb.filter(F.action == "export_audience"))
async def cb_bot_export_audience(
    callback: CallbackQuery, callback_data: BotCb, pool: asyncpg.Pool
) -> None:
    """Export all bot users as a CSV file. Available from STARTER plan."""
    if not await require_plan(pool, callback.from_user.id, "starter"):
        await callback.answer()
        await callback.message.edit_text(
            locked_text("Экспорт аудитории", "starter"),
            parse_mode="HTML",
            reply_markup=subscription_locked_markup("starter"),
        )
        return

    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return

    await callback.answer("⏳ Генерирую CSV…")

    rows = await db.get_audience_full(pool, callback_data.bot_id)
    if not rows:
        await callback.message.answer("📤 Аудитория пуста — нечего экспортировать.")
        return

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "user_id",
            "username",
            "first_name",
            "last_name",
            "language_code",
            "first_seen",
            "last_seen",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r["user_id"],
                r["username"] or "",
                r["first_name"] or "",
                r["last_name"] or "",
                r["language_code"] or "",
                r["first_seen"].strftime("%Y-%m-%d %H:%M:%S"),
                r["last_seen"].strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    label = f"@{row['username']}" if row["username"] else row["first_name"]
    safe_label = row["username"] or str(callback_data.bot_id)
    filename = f"audience_{safe_label}.csv"
    content = buf.getvalue().encode("utf-8-sig")

    await callback.message.answer_document(
        BufferedInputFile(content, filename=filename),
        caption=f"📤 <b>Экспорт аудитории {label}</b>\n\nЗаписей: <b>{len(rows)}</b>",
        parse_mode="HTML",
    )


# ── Block / unblock user ───────────────────────────────────────────────────


@router.callback_query(AudCb.filter(F.action.in_({"block_user", "unblock_user"})))
async def cb_block_user(
    callback: CallbackQuery, callback_data: AudCb, pool: asyncpg.Pool
) -> None:

    blocked = callback_data.action == "block_user"
    row = await db.get_bot(pool, callback_data.bot_id, callback.from_user.id)
    if not row:
        await callback.answer("Бот не найден.", show_alert=True)
        return
    await db.block_user(pool, callback_data.bot_id, callback_data.target_id, blocked)
    user = await db.get_user_by_id(pool, callback_data.bot_id, callback_data.target_id)
    if not user:
        await callback.answer("Пользователь не найден.", show_alert=True)
        return
    _uname = user.get("username")
    _fname = (user.get("first_name") or "").strip()
    _lname = (user.get("last_name") or "").strip()
    _full_name = f"{_fname} {_lname}".strip() or None
    u_label = f"@{_uname}" if _uname else (_full_name or str(user["user_id"]))
    action_text = "заблокирован" if blocked else "разблокирован"
    lang = user.get("language_code") or "—"
    phone_line = (
        f"\n📱 Телефон: <code>{user['phone']}</code>" if user.get("phone") else ""
    )
    await callback.message.edit_text(
        f"👤 <b>Пользователь {u_label}</b>\n"
        f"ID: <code>{user['user_id']}</code>\n"
        + (f"Имя: {_full_name}\n" if _full_name else "")
        + phone_line
        + f"\nЯзык: {lang}\n"
        f"Статус: {'🚫 Заблокирован' if blocked else '✅ Активен'}",
        parse_mode="HTML",
        reply_markup=user_profile_menu(
            callback_data.bot_id, callback_data.target_id, blocked
        ),
    )
    await callback.answer(f"✅ Пользователь {action_text}.")
